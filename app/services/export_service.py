"""CSV/XLSX report export service for Phase 3.

The project is still in a staged migration: core legacy data may live in the
SQLite database while new admin/ticket/plan/broadcast data lives in PostgreSQL.
This service exports from both sources so reports are useful during the bridge
period and after the full migration.
"""

from __future__ import annotations

import csv
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from app.config import settings
from app.database import session_scope
from app.models import (
    Admin,
    AdminLog,
    BroadcastCampaign,
    BroadcastRecipient,
    CatalogPlan,
    Coupon,
    DataAddonPackageDB,
    Order,
    Referral,
    Service,
    TextTemplate,
    Ticket,
    TicketMessage,
    User,
    WalletTransaction,
)


@dataclass(frozen=True)
class ReportSpec:
    key: str
    title: str
    source: str  # sqlite | pg | virtual
    table: str | None = None
    model: type[Any] | None = None


REPORT_SPECS: dict[str, ReportSpec] = {
    "users": ReportSpec("users", "کاربران", "sqlite", "users", User),
    "services": ReportSpec("services", "سرویس‌ها", "sqlite", "services", Service),
    "orders": ReportSpec("orders", "سفارش‌ها", "sqlite", "orders", Order),
    "wallet": ReportSpec("wallet", "تراکنش‌های کیف پول", "sqlite", "wallet_transactions", WalletTransaction),
    "payment_cards": ReportSpec("payment_cards", "کارت‌های پرداخت", "sqlite", "payment_cards"),
    "payment_receipts": ReportSpec("payment_receipts", "رسیدهای کارت‌به‌کارت", "sqlite", "payment_receipts"),
    "financial_ledger": ReportSpec("financial_ledger", "گردش مالی کامل", "virtual"),
    "finance_summary": ReportSpec("finance_summary", "خلاصه مالی", "virtual"),
    "referrals": ReportSpec("referrals", "رفرال‌ها", "sqlite", "referrals", Referral),
    "tickets": ReportSpec("tickets", "تیکت‌ها", "pg", "tickets", Ticket),
    "ticket_messages": ReportSpec("ticket_messages", "پیام‌های تیکت", "pg", "ticket_messages", TicketMessage),
    "admins": ReportSpec("admins", "ادمین‌ها", "pg", "admins", Admin),
    "admin_logs": ReportSpec("admin_logs", "لاگ ادمین‌ها", "sqlite", "admin_logs", AdminLog),
    "coupons": ReportSpec("coupons", "کدهای تخفیف", "sqlite", "coupons", Coupon),
    "plans": ReportSpec("plans", "پلن‌ها", "pg", "plans", CatalogPlan),
    "addons": ReportSpec("addons", "بسته‌های افزایش حجم", "pg", "data_addon_packages", DataAddonPackageDB),
    "texts": ReportSpec("texts", "متن‌های قابل ویرایش", "pg", "text_templates", TextTemplate),
    "broadcasts": ReportSpec("broadcasts", "پیام‌های همگانی", "pg", "broadcast_campaigns", BroadcastCampaign),
    "broadcast_recipients": ReportSpec("broadcast_recipients", "گیرندگان پیام همگانی", "pg", "broadcast_recipients", BroadcastRecipient),
    "usage": ReportSpec("usage", "گزارش استفاده", "virtual"),
}


def _now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def _serialize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _sqlite_path() -> Path:
    return Path(settings.database_path)


def _sqlite_connect() -> sqlite3.Connection | None:
    path = _sqlite_path()
    if not path.exists():
        return None
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def _sqlite_table_exists(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return row is not None


def _sqlite_rows(table: str, limit: int | None = None) -> list[dict[str, Any]]:
    conn = _sqlite_connect()
    if conn is None:
        return []
    try:
        if not _sqlite_table_exists(conn, table):
            return []
        sql = f"SELECT * FROM {table} ORDER BY 1 DESC"
        params: tuple[Any, ...] = ()
        if limit:
            sql += " LIMIT ?"
            params = (limit,)
        return [dict(row) for row in conn.execute(sql, params).fetchall()]
    finally:
        conn.close()


async def _pg_rows(model: type[Any], limit: int | None = None) -> list[dict[str, Any]]:
    async with session_scope() as session:
        stmt = select(model)
        if hasattr(model, "id"):
            stmt = stmt.order_by(getattr(model, "id").desc())
        elif hasattr(model, "created_at"):
            stmt = stmt.order_by(getattr(model, "created_at").desc())
        if limit:
            stmt = stmt.limit(limit)
        result = await session.execute(stmt)
        rows = result.scalars().all()
        output: list[dict[str, Any]] = []
        for obj in rows:
            data: dict[str, Any] = {}
            for col in obj.__table__.columns:
                data[col.name] = _serialize(getattr(obj, col.name))
            output.append(data)
        return output


def _sqlite_scalar(query: str, params: tuple[Any, ...] = ()) -> int | float:
    conn = _sqlite_connect()
    if conn is None:
        return 0
    try:
        row = conn.execute(query, params).fetchone()
        if not row or row[0] is None:
            return 0
        return row[0]
    except Exception:
        return 0
    finally:
        conn.close()


def _sqlite_group_rows(query: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    conn = _sqlite_connect()
    if conn is None:
        return []
    try:
        return [dict(row) for row in conn.execute(query, params).fetchall()]
    except Exception:
        return []
    finally:
        conn.close()


def _financial_summary_rows() -> list[dict[str, Any]]:
    created_at = datetime.now(timezone.utc).isoformat()
    rows: list[dict[str, Any]] = [
        {"section": "orders", "metric": "orders_total", "title": "تعداد کل سفارش‌ها", "value": _sqlite_scalar("SELECT COUNT(*) FROM orders"), "created_at": created_at},
        {"section": "orders", "metric": "orders_paid", "title": "سفارش‌های پرداخت‌شده", "value": _sqlite_scalar("SELECT COUNT(*) FROM orders WHERE status = 'paid'"), "created_at": created_at},
        {"section": "orders", "metric": "orders_pending", "title": "سفارش‌های در انتظار پرداخت", "value": _sqlite_scalar("SELECT COUNT(*) FROM orders WHERE status = 'pending'"), "created_at": created_at},
        {"section": "orders", "metric": "orders_receipt_pending", "title": "رسیدهای در انتظار بررسی", "value": _sqlite_scalar("SELECT COUNT(*) FROM orders WHERE status = 'receipt_pending'"), "created_at": created_at},
        {"section": "orders", "metric": "orders_rejected", "title": "پرداخت‌های ردشده", "value": _sqlite_scalar("SELECT COUNT(*) FROM orders WHERE status = 'payment_rejected'"), "created_at": created_at},
        {"section": "orders", "metric": "gross_paid", "title": "مبلغ خام سفارش‌های پرداخت‌شده", "value": _sqlite_scalar("SELECT COALESCE(SUM(amount), 0) FROM orders WHERE status = 'paid'"), "created_at": created_at},
        {"section": "orders", "metric": "discount_paid", "title": "جمع تخفیف سفارش‌های پرداخت‌شده", "value": _sqlite_scalar("SELECT COALESCE(SUM(discount_amount), 0) FROM orders WHERE status = 'paid'"), "created_at": created_at},
        {"section": "orders", "metric": "net_paid", "title": "خالص فروش پرداخت‌شده", "value": _sqlite_scalar("SELECT COALESCE(SUM(amount - discount_amount), 0) FROM orders WHERE status = 'paid'"), "created_at": created_at},
        {"section": "wallet", "metric": "wallet_debits", "title": "برداشت از کیف پول", "value": _sqlite_scalar("SELECT COALESCE(SUM(ABS(amount)), 0) FROM wallet_transactions WHERE amount < 0"), "created_at": created_at},
        {"section": "wallet", "metric": "wallet_credits", "title": "واریز/بازگشت به کیف پول", "value": _sqlite_scalar("SELECT COALESCE(SUM(amount), 0) FROM wallet_transactions WHERE amount > 0"), "created_at": created_at},
        {"section": "receipts", "metric": "receipts_total", "title": "تعداد رسیدهای کارت‌به‌کارت", "value": _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts"), "created_at": created_at},
        {"section": "receipts", "metric": "receipts_pending", "title": "رسیدهای در انتظار", "value": _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts WHERE status = 'receipt_pending'"), "created_at": created_at},
        {"section": "receipts", "metric": "receipts_approved", "title": "رسیدهای تأییدشده", "value": _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts WHERE status = 'approved'"), "created_at": created_at},
        {"section": "receipts", "metric": "receipts_rejected", "title": "رسیدهای ردشده", "value": _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts WHERE status = 'rejected'"), "created_at": created_at},
    ]
    for row in _sqlite_group_rows("SELECT COALESCE(payment_method, 'unknown') AS method, COUNT(*) AS count, COALESCE(SUM(amount - discount_amount), 0) AS net_amount FROM orders WHERE status = 'paid' GROUP BY COALESCE(payment_method, 'unknown') ORDER BY net_amount DESC"):
        rows.append({"section": "paid_by_method", "metric": row.get("method"), "title": f"فروش پرداخت‌شده با {row.get('method')}", "value": row.get("net_amount"), "count": row.get("count"), "created_at": created_at})
    for row in _sqlite_group_rows("SELECT COALESCE(type, 'unknown') AS tx_type, COUNT(*) AS count, COALESCE(SUM(amount), 0) AS amount FROM wallet_transactions GROUP BY COALESCE(type, 'unknown') ORDER BY amount DESC"):
        rows.append({"section": "wallet_by_type", "metric": row.get("tx_type"), "title": f"کیف پول / {row.get('tx_type')}", "value": row.get("amount"), "count": row.get("count"), "created_at": created_at})
    return rows


def _financial_ledger_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in _sqlite_rows("orders"):
        rows.append({
            "source": "orders",
            "event_id": row.get("id"),
            "created_at": row.get("created_at"),
            "user_telegram_id": row.get("user_telegram_id"),
            "status": row.get("status"),
            "method": row.get("payment_method"),
            "gross_amount": row.get("amount"),
            "discount_amount": row.get("discount_amount"),
            "net_amount": (int(row.get("amount") or 0) - int(row.get("discount_amount") or 0)),
            "wallet_used": row.get("wallet_used"),
            "coupon_code": row.get("coupon_code"),
            "receipt_id": row.get("receipt_id"),
            "service_id": row.get("service_id"),
            "description": row.get("admin_note") or row.get("plan_key"),
        })
    for row in _sqlite_rows("wallet_transactions"):
        rows.append({
            "source": "wallet_transactions",
            "event_id": row.get("id"),
            "created_at": row.get("created_at"),
            "user_telegram_id": row.get("user_telegram_id"),
            "status": "posted",
            "method": "wallet",
            "gross_amount": row.get("amount"),
            "discount_amount": 0,
            "net_amount": row.get("amount"),
            "wallet_used": None,
            "coupon_code": None,
            "receipt_id": None,
            "service_id": None,
            "description": row.get("description") or row.get("type"),
        })
    for row in _sqlite_rows("payment_receipts"):
        rows.append({
            "source": "payment_receipts",
            "event_id": row.get("id"),
            "created_at": row.get("created_at"),
            "user_telegram_id": row.get("user_telegram_id"),
            "status": row.get("status"),
            "method": "card_to_card",
            "gross_amount": row.get("amount"),
            "discount_amount": 0,
            "net_amount": row.get("amount"),
            "wallet_used": None,
            "coupon_code": None,
            "receipt_id": row.get("id"),
            "service_id": None,
            "description": row.get("admin_note") or row.get("receipt_caption") or f"order_id={row.get('order_id')}",
        })
    return sorted(rows, key=lambda item: str(item.get("created_at") or ""), reverse=True)


async def usage_summary_rows() -> list[dict[str, Any]]:
    """Return a compact bot usage report spanning SQLite and PostgreSQL."""
    sqlite_metrics = [
        ("users_total", "تعداد کل کاربران", _sqlite_scalar("SELECT COUNT(*) FROM users")),
        ("users_active", "کاربران فعال", _sqlite_scalar("SELECT COUNT(*) FROM users WHERE COALESCE(status, 'active') = 'active'")),
        ("services_total", "تعداد کل سرویس‌ها", _sqlite_scalar("SELECT COUNT(*) FROM services")),
        ("services_active", "سرویس‌های فعال", _sqlite_scalar("SELECT COUNT(*) FROM services WHERE status = 'active'")),
        ("services_suspended", "سرویس‌های غیرفعال/قفل", _sqlite_scalar("SELECT COUNT(*) FROM services WHERE status != 'active'")),
        ("data_total_gb", "حجم کل سرویس‌ها GB", _sqlite_scalar("SELECT COALESCE(SUM(data_gb), 0) FROM services")),
        ("data_used_gb", "حجم مصرف‌شده GB", round(float(_sqlite_scalar("SELECT COALESCE(SUM(data_used_mb), 0) FROM services")) / 1024, 2)),
        ("orders_total", "تعداد سفارش‌ها", _sqlite_scalar("SELECT COUNT(*) FROM orders")),
        ("orders_paid", "سفارش‌های پرداخت‌شده", _sqlite_scalar("SELECT COUNT(*) FROM orders WHERE status = 'paid'")),
        ("sales_total", "فروش کل پرداخت‌شده", _sqlite_scalar("SELECT COALESCE(SUM(amount - discount_amount), 0) FROM orders WHERE status = 'paid'")),
        ("receipts_pending", "رسیدهای کارت‌به‌کارت در انتظار", _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts WHERE status = 'receipt_pending'")),
        ("receipts_approved", "رسیدهای تأییدشده", _sqlite_scalar("SELECT COUNT(*) FROM payment_receipts WHERE status = 'approved'")),
        ("wallet_balance_total", "جمع موجودی کیف پول کاربران", _sqlite_scalar("SELECT COALESCE(SUM(wallet_balance), 0) FROM users")),
    ]
    pg_metrics: list[tuple[str, str, Any]] = []
    async with session_scope() as session:
        for key, title, model, where_status in [
            ("tickets_total", "تعداد کل تیکت‌ها", Ticket, None),
            ("tickets_open", "تیکت‌های باز", Ticket, "open"),
            ("tickets_waiting_admin", "تیکت‌های منتظر ادمین", Ticket, "waiting_admin"),
            ("tickets_closed", "تیکت‌های بسته", Ticket, "closed"),
            ("broadcasts_total", "کمپین‌های پیام همگانی", BroadcastCampaign, None),
            ("plans_active", "پلن‌های فعال", CatalogPlan, None),
        ]:
            stmt = select(func.count()).select_from(model)
            if where_status:
                stmt = stmt.where(model.status == where_status)
            elif key == "plans_active":
                stmt = stmt.where(CatalogPlan.is_active.is_(True))
            result = await session.execute(stmt)
            pg_metrics.append((key, title, int(result.scalar() or 0)))
    created_at = datetime.now(timezone.utc).isoformat()
    return [
        {"key": key, "title": title, "value": value, "created_at": created_at}
        for key, title, value in [*sqlite_metrics, *pg_metrics]
    ]


async def collect_report_rows(report_key: str, limit: int | None = None) -> tuple[ReportSpec, list[dict[str, Any]]]:
    spec = REPORT_SPECS[report_key]
    if spec.source == "virtual":
        if report_key == "usage":
            return spec, await usage_summary_rows()
        if report_key == "finance_summary":
            return spec, _financial_summary_rows()
        if report_key == "financial_ledger":
            return spec, _financial_ledger_rows()
        return spec, []
    if spec.source == "sqlite":
        rows = _sqlite_rows(spec.table or "", limit=limit)
        # During/after migration, PostgreSQL may be the source of truth. Fall back to PG if SQLite is empty.
        if not rows and spec.model is not None:
            rows = await _pg_rows(spec.model, limit=limit)
        return spec, rows
    if spec.model is None:
        return spec, []
    return spec, await _pg_rows(spec.model, limit=limit)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = sorted({key for row in rows for key in row.keys()}) if rows else ["empty"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        if rows:
            writer.writerows(rows)


def _write_xlsx(path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    headers = sorted({key for row in rows for key in row.keys()}) if rows else ["empty"]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="E8F0FE")
    ws.append([f"Generated at: {datetime.now(timezone.utc).isoformat()}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.alignment = Alignment(horizontal="center")
    if rows:
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    else:
        ws.append(["No rows"])
    for idx, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows[:200]]
        width = min(max(len(v) for v in values) + 2, 48)
        ws.column_dimensions[get_column_letter(idx)].width = max(width, 12)
    wb.save(path)


async def build_report_file(report_key: str, file_format: str, output_dir: str | Path = "/tmp/howtosee_reports") -> tuple[Path, int, str]:
    spec, rows = await collect_report_rows(report_key)
    ext = "xlsx" if file_format == "xlsx" else "csv"
    path = Path(output_dir) / f"report-{report_key}-{_now_stamp()}.{ext}"
    if ext == "xlsx":
        _write_xlsx(path, spec.title, rows)
    else:
        _write_csv(path, rows)
    return path, len(rows), spec.title


async def build_all_reports_zip(output_dir: str | Path = "/tmp/howtosee_reports") -> Path:
    import zipfile

    base = Path(output_dir)
    base.mkdir(parents=True, exist_ok=True)
    zip_path = base / f"reports-all-{_now_stamp()}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for key in REPORT_SPECS:
            path, _, _ = await build_report_file(key, "csv", base / "parts")
            zf.write(path, arcname=f"{key}.csv")
    return zip_path
